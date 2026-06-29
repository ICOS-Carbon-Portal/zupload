# Standard library imports.
from pathlib import Path
from pprint import pprint
from typing import Any, Literal
import json
import ast
import re
from datetime import datetime as dt
from urllib.parse import urlparse, urlunparse
# Related third party imports.
import numpy as np
import pandas as pd
import pyproj
import xarray
from halo import Halo
from openpyxl import Workbook, load_workbook
from openpyxl.pivot.fields import Boolean
from pandas import Series
import typer
import requests
# Local application/library specific imports.
from zupload.utils import (
    calculate_hashsum,
    get_conf,
    get_cookie_jar,
    write_json,
    get_prev_by_name
)
from zupload.constants.envri import EnvriConfig
from zupload.constants.object_specs import ALL_OBJECT_SPECS
from zupload.constants.organizations import (
    ORG_DISPLAY_NAMES,
    ORG_LU_CITIES,
)
from zupload.constants.stations import CITIES_FOR_STATION
from zupload.constants.upload_descriptions import CITIES_UPLOAD_DESCRIPTIONS
from zupload.logs import RunLogger
from zupload.validation import validate_columns, validate_dataframe


app = typer.Typer(help='Upload data & metadata to the specific portal.')


def _portal_display_name(envri_name: str) -> str:
    return {
        'ICOSCities': 'cities',
        'ICOS': 'icos',
        'SITES': 'sites',
    }.get(envri_name, envri_name.lower())


def _resolve_spreadsheet(spreadsheet: str | None) -> Path:
    if spreadsheet is None:
        matches = list(Path.cwd().glob('*.xlsx'))
        if not matches:
            typer.echo('No .xlsx files found in current directory.')
            raise typer.Exit(code=1)
        if len(matches) > 1:
            typer.echo('More than one spreadsheet found in current directory:')
            for match in matches:
                typer.echo(f'- {match.name}')
            typer.echo('Please rerun by explicitly passing the spreadsheet path, for example:')
            typer.echo('zupload ./your_spreadsheet.xlsx [options]')
            raise typer.Exit(code=1)
        return matches[0]
    return Path(spreadsheet)


def _to_landing_uri(pid: str) -> str:
    pid = pid.strip()
    if pid.startswith('http://') or pid.startswith('https://'):
        return pid
    return f'https://meta.icos-cp.eu/objects/{pid}'


def _select_rows(df, rows_value, flag='--rows'):
    """Slice df to the given upload_meta sheet row spec ("5" or "5-12", inclusive). Returns the sliced df."""
    value = rows_value.strip()
    if '-' in value:
        parts = value.split('-')
        if len(parts) != 2 or not parts[0] or not parts[1]:
            typer.echo(
                f'Invalid {flag} value "{rows_value}". '
                'Expected an integer like "5" or a range like "5-12".'
            )
            raise typer.Exit(code=1)
        try:
            start = int(parts[0])
            end = int(parts[1])
        except ValueError:
            typer.echo(
                f'Invalid {flag} value "{rows_value}". '
                'Expected an integer like "5" or a range like "5-12".'
            )
            raise typer.Exit(code=1)
        if start > end:
            typer.echo(
                f'Invalid {flag} range "{rows_value}": '
                'start must be <= end.'
            )
            raise typer.Exit(code=1)
    else:
        try:
            start = int(value)
        except ValueError:
            typer.echo(
                f'Invalid {flag} value "{rows_value}". '
                'Expected an integer like "5" or a range like "5-12".'
            )
            raise typer.Exit(code=1)
        end = start
    if start < 2 or end > (len(df) + 1):
        typer.echo(
            f'Invalid {flag} range "{rows_value}". '
            f'Expected 2..{len(df) + 1} for upload_meta.'
        )
        raise typer.Exit(code=1)
    return df.iloc[start - 2 : end - 1]


@app.command()
def fetch(
        pid: str = typer.Argument(
            ...,
            metavar='PID',
            help='Landing page URL, object id, or object hash.'
        )
):
    """Fetch and print dataset metadata from the ICOS portal by landing page URL, object id, or hash."""
    landing_uri = _to_landing_uri(pid)
    typer.echo(f'Fetching metadata for: {landing_uri}')
    resp = requests.get(
        'https://meta.icos-cp.eu/dtodownload',
        params={'uri': landing_uri},
    )
    if resp.status_code != 200:
        typer.echo(f'Fetch failed ({resp.status_code})')
        typer.echo(resp.text)
        raise typer.Exit(code=1)
    try:
        payload = resp.json()
    except ValueError:
        try:
            payload = json.loads(resp.text)
        except json.JSONDecodeError:
            payload = ast.literal_eval(resp.text)
    pprint(payload)

@app.command()
def validate(
        spreadsheet: str | None = typer.Option(
            None,
            '--spreadsheet',
            help='Spreadsheet to use (.xlsx). Auto-detected if the current directory has a single .xlsx.'
        ),
        rows: str | None = typer.Option(
            None,
            '--rows',
            help='Validate a single row or a contiguous range by upload_meta sheet row numbers, e.g. "5" or "5-12" (inclusive on both ends).'
        ),
        data_dir: str | None = typer.Option(
            None,
            '--data-dir',
            help='Locate each data file by name under this directory, fill in fileLocation and hashSum where resolvable, and write a new <name>.filled.xlsx (the original spreadsheet is left unchanged).'
        )
):
    """Check upload_meta rows for metadata problems without uploading or changing the spreadsheet."""
    spreadsheet = _resolve_spreadsheet(spreadsheet)
    df = pd.read_excel(spreadsheet, sheet_name='upload_meta')
    schema_issues = validate_columns(df)
    filled_path = None
    resolved = 0
    ambiguous = 0
    not_found = 0
    if data_dir is not None:
        data_root = Path(data_dir)
        if not data_root.is_dir():
            typer.echo(f'--data-dir not found or not a directory: {data_dir}')
            raise typer.Exit(code=1)
        name_to_paths: dict[str, list[Path]] = {}
        for candidate in data_root.rglob('*'):
            if candidate.is_file():
                name_to_paths.setdefault(candidate.name, []).append(candidate)
        for col_name in ('fileLocation', 'hashSum'):
            if col_name in df.columns:
                df[col_name] = df[col_name].astype(object)
        updates = []
        for idx, row in df.iterrows():
            file_name = row.get('fileName')
            if pd.isna(file_name) or not str(file_name).strip():
                continue
            file_name = str(file_name).strip()
            target = None
            file_location = row.get('fileLocation')
            if not pd.isna(file_location) and str(file_location).strip():
                candidate = Path(str(file_location).strip()) / file_name
                if candidate.exists():
                    target = candidate
            if target is None:
                matches = name_to_paths.get(file_name, [])
                if len(matches) == 1:
                    target = matches[0]
                elif len(matches) > 1:
                    ambiguous += 1
                    continue
                else:
                    not_found += 1
                    continue
            new_loc = str(target.parent.resolve())
            df.at[idx, 'fileLocation'] = new_loc
            updates.append((idx + 2, 'fileLocation', new_loc))
            hash_sum = row.get('hashSum')
            if pd.isna(hash_sum) or not str(hash_sum).strip():
                new_hash = calculate_hashsum(file_path=target)
                df.at[idx, 'hashSum'] = new_hash
                updates.append((idx + 2, 'hashSum', new_hash))
            resolved += 1
        wb = load_workbook(spreadsheet)
        ws = wb['upload_meta']
        headers = {cell.value: i for i, cell in enumerate(ws[1], start=1)}
        col_index = {}
        for col_name in ('fileLocation', 'hashSum'):
            c = headers.get(col_name)
            if c is None:
                c = ws.max_column + 1
                ws.cell(row=1, column=c).value = col_name
            col_index[col_name] = c
        for sheet_row, col_name, value in updates:
            ws.cell(row=sheet_row, column=col_index[col_name]).value = value
        filled_path = spreadsheet.with_name(f'{spreadsheet.stem}.filled.xlsx')
        wb.save(filled_path)
    if rows is not None:
        df = _select_rows(df, rows)
    results = validate_dataframe(df)
    total = len(results)
    rows_with_errors = 0
    rows_with_warnings = 0
    ok_rows = 0
    for result in results:
        typer.echo(f'Row {result["row"]}: {result["fileName"]}')
        issues = result['issues']
        if not issues:
            typer.echo('  ok')
            ok_rows += 1
            continue
        has_error = any(severity == 'error' for severity, _ in issues)
        has_warning = any(severity == 'warning' for severity, _ in issues)
        if has_error:
            rows_with_errors += 1
        if has_warning:
            rows_with_warnings += 1
        for severity, message in issues:
            typer.echo(f'  {severity}: {message}')
    if schema_issues:
        typer.echo('Schema check:')
        for severity, message in schema_issues:
            typer.echo(f'  {severity}: {message}')
    else:
        typer.echo('Schema check: ok')
    schema_errors = sum(1 for severity, _ in schema_issues if severity == 'error')
    schema_warnings = sum(1 for severity, _ in schema_issues if severity == 'warning')
    typer.echo(
        f'{total} rows checked - {rows_with_errors} with errors, '
        f'{rows_with_warnings} with warnings, {ok_rows} ok; '
        f'{schema_errors} schema errors, {schema_warnings} schema warnings'
    )
    if filled_path is not None:
        typer.echo(f'Filled spreadsheet written to {filled_path}')
        typer.echo(
            f'Data resolution: {resolved} filled, {ambiguous} ambiguous '
            f'(same filename in multiple places), {not_found} not found in --data-dir'
        )


@app.callback(invoke_without_command=True)
def main(
        ctx: typer.Context,
        spreadsheet: str | None = typer.Option(
            None,
            '--spreadsheet',
            help='Spreadsheet to use (.xlsx). Auto-detected if the current directory has a single .xlsx.'
        ),
        extract_json: bool = typer.Option(
            False,
            help="Write each row's metadata JSON next to its data file and skip the upload."
        ),
        upload: bool = typer.Option(
            True,
            help='Upload to the portal (default). Use --no-upload for a dry run that builds metadata without uploading or writing files.'
        ),
        metadata_only: bool = typer.Option(
            False,
            '--metadata-only',
            help='Upload metadata only and skip data file upload.'
        ),
        rows: str | None = typer.Option(
            None,
            '--rows',
            help='Upload a single row or a contiguous range by upload_meta sheet row numbers, e.g. "5" or "5-12" (inclusive on both ends).'
        )
):
    run_logger: RunLogger | None = None
    if ctx.invoked_subcommand:
        return
    try:
        if extract_json:
            upload = False
        spreadsheet = _resolve_spreadsheet(spreadsheet)
        run_logger = RunLogger.start(spreadsheet=spreadsheet)
        envri_conf = get_conf(file_path=spreadsheet)
        if upload:
            typer.echo(f'Using portal: {_portal_display_name(envri_conf.envri)}')
        wb = load_workbook(spreadsheet)
        ws = wb['upload_meta']
        headers = {cell.value: i for i, cell in enumerate(ws[1], start=1)}
        data_url_col = headers.get('dataUploadUrl')
        if data_url_col is None:
            data_url_col = ws.max_column + 1
            ws.cell(row=1, column=data_url_col).value = 'dataUploadUrl'
        landing_col = headers.get('landingPageURI')
        if landing_col is None:
            landing_col = ws.max_column + 1
            ws.cell(row=1, column=landing_col).value = 'landingPageURI'
        df = pd.read_excel(spreadsheet, sheet_name='upload_meta')
        if rows is not None:
            df = _select_rows(df, rows)
        for idx, row in df.iterrows():
            typer.echo(f'Row {idx + 2}: {row["fileName"]}')
            meta_json = make_json(meta=row)
            if upload:
                data_url, landing_url = upload_meta(meta_json=meta_json, envri_conf=envri_conf)
                ws.cell(row=idx + 2, column=data_url_col).value = data_url
                ws.cell(row=idx + 2, column=landing_col).value = landing_url
                wb.save(spreadsheet)
                if not metadata_only:
                    upload_data(file_path=Path(row['fileLocation']) / row['fileName'], data_url=data_url)
            elif extract_json:
                p = Path(row['fileLocation']) / row['fileName']
                json_path = p.with_suffix('.json')
                write_json(file=json_path, content=meta_json)
                pprint(meta_json)
                typer.echo(f'JSON written to {json_path}')
            else:
                pprint(meta_json)
                typer.echo('Dry run: metadata built, nothing uploaded or written (use --extract-json to write JSON).')
        if upload:
            typer.echo(f'Updated upload URLs in {spreadsheet}')
    except Exception as e:
        if run_logger is not None:
            run_logger.finish(status='error', error=str(e))
        raise
    if run_logger is not None:
        run_logger.finish(status='ok')

@app.command()
def generate(
        directory: Path = typer.Argument(
            Path('.'),
            exists=True,
            file_okay=False,
            dir_okay=True,
            help='Directory of NetCDF data files to scan (defaults to the current directory).'
        ),
        output: Path = typer.Option(
            Path('upload_meta.xlsx'),
            '--output',
            '-o',
            help='Spreadsheet to create or update (default: upload_meta.xlsx).'
        ),
        portal: str = typer.Option(
            'icos',
            '--portal',
            help='Target portal: icos, sites, or cities.'
        ),
        spec_label: str = typer.Option(
            'Non-standard spatial product',
            '--spec-label',
            help='Object specification label; must match a known spec in constants/object_specs.py.'
        ),
        hash_mode: str = typer.Option(
            'auto',
            '--hash-mode',
            help=('Hash strategy: auto=reuse existing hashes, compute missing; '
                  'reuse=only reuse existing; recompute=compute all hashes again.')
        ),
        prev_mode: str = typer.Option(
            'auto',
            '--prev-mode',
            help=('Previous-version strategy: auto=reuse existing values, query missing; '
                  'reuse=only reuse existing; recompute=query all via SPARQL; skip=leave empty.')
        ),
        description_key: Literal['paris', 'munich', 'zurich'] = typer.Option(
            ...,
            '--description-key',
            help='Predefined description key from constants/upload_descriptions.py.'
        ),
        update_columns: str = typer.Option(
            '',
            '--update-columns',
            help=(
                'Comma-separated columns to update in-place from static metadata '
                '(e.g. comment,forStation,description).'
            )
        ),
):
    """Scaffold an upload_meta.xlsx for ICOS Cities footprint NetCDF files in a directory."""
    run_logger = RunLogger.start(spreadsheet=output)
    typer.echo(f'Scanning input directory: {directory}')
    try:
        if spec_label not in ALL_OBJECT_SPECS:
            typer.echo('Invalid spec label. Available labels:')
            for label in sorted(ALL_OBJECT_SPECS):
                typer.echo(f'- {label}')
            raise typer.Exit(code=1)
        if hash_mode not in {'auto', 'reuse', 'recompute'}:
            typer.echo('Invalid hash mode. Use one of: auto, reuse, recompute')
            raise typer.Exit(code=1)
        if prev_mode not in {'auto', 'reuse', 'recompute', 'skip'}:
            typer.echo(
                'Invalid prev mode. Use one of: auto, reuse, recompute, skip'
            )
            raise typer.Exit(code=1)
        portal_input = portal.strip().lower()
        allowed_portals = {'icos', 'sites', 'cities'}
        if portal_input not in allowed_portals:
            typer.echo('Invalid portal. Use one of: icos, sites, cities')
            raise typer.Exit(code=1)
        portal_aliases = {'cities': 'icoscities'}
        portal_norm = portal_aliases.get(portal_input, portal_input)
        description = CITIES_UPLOAD_DESCRIPTIONS[str(description_key)]
        for_station = CITIES_FOR_STATION[str(description_key)]
        static_meta = build_static_cities_meta(
            description=description,
            for_station=for_station
        )
        typer.echo(
            f'Options: portal={portal_input}, type="{spec_label}", '
            f'hash-mode={hash_mode}, prev-mode={prev_mode}, description-key={description_key}'
        )
        if update_columns:
            if not output.exists():
                typer.echo(f'Cannot update columns: output file does not exist: {output}')
                raise typer.Exit(code=1)
            alias_map = {'description': 'abstract/description '}
            selected_input = [c.strip() for c in update_columns.split(',') if c.strip()]
            selected = [alias_map.get(c, c) for c in selected_input]
            if not selected:
                typer.echo('No columns provided in --update-columns.')
                raise typer.Exit(code=1)
            wb = load_workbook(output)
            if 'upload_meta' not in wb.sheetnames:
                typer.echo('Cannot update columns: sheet "upload_meta" not found.')
                raise typer.Exit(code=1)
            ws = wb['upload_meta']
            headers = {cell.value: i for i, cell in enumerate(ws[1], start=1)}
            missing = [c for c in selected if c not in headers]
            if missing:
                typer.echo(f'Columns not found in upload_meta: {", ".join(missing)}')
                raise typer.Exit(code=1)
            unsupported = [c for c in selected if c not in static_meta]
            if unsupported:
                typer.echo(
                    f'Columns are not static-update columns: {", ".join(unsupported)}'
                )
                raise typer.Exit(code=1)
            for row in range(2, ws.max_row + 1):
                for col in selected:
                    ws.cell(row=row, column=headers[col]).value = static_meta[col]
            wb.save(output)
            typer.echo(f'Updated columns in {output}: {", ".join(selected)}')
            run_logger.finish(status='ok')
            return
        object_spec = ALL_OBJECT_SPECS[spec_label]
        files = sorted(p.name for p in directory.iterdir() if p.is_file())
        typer.echo(f'Found {len(files)} files')
        existing_hashes: dict[str, str] = {}
        existing_prev: dict[str, str] = {}
        should_reuse = hash_mode in {'auto', 'reuse'}
        should_reuse_prev = prev_mode in {'auto', 'reuse'}
        if output.exists() and (should_reuse or should_reuse_prev):
            typer.echo(f'Loading reusable values from existing output: {output}')
            wb_prev = load_workbook(output, data_only=True)
            if 'upload_meta' in wb_prev.sheetnames:
                ws_prev = wb_prev['upload_meta']
                headers = {cell.value: i for i, cell in enumerate(ws_prev[1], start=1)}
                loc_idx = headers.get('fileLocation')
                hash_idx = headers.get('hashSum')
                prev_idx = headers.get('isNextVersionOf')
                if loc_idx and hash_idx and should_reuse:
                    for row in range(2, ws_prev.max_row + 1):
                        file_location = ws_prev.cell(row=row, column=loc_idx).value
                        hash_sum = ws_prev.cell(row=row, column=hash_idx).value
                        if file_location and hash_sum:
                            existing_hashes[str(file_location)] = str(hash_sum)
                if loc_idx and prev_idx and should_reuse_prev:
                    for row in range(2, ws_prev.max_row + 1):
                        file_location = ws_prev.cell(row=row, column=loc_idx).value
                        prev_version = ws_prev.cell(row=row, column=prev_idx).value
                        if file_location and prev_version:
                            existing_prev[str(file_location)] = str(prev_version)

        rows = []
        hash_reused = 0
        hash_computed = 0
        prev_reused = 0
        prev_queried = 0
        prev_empty = 0
        spinner = Halo(text='Extracting metadata and building rows...', spinner='dots')
        spinner.start()
        try:
            for name in files:
                file_path = directory / name
                file_location = str(file_path.parent.resolve())
                hash_sum = None if hash_mode == 'recompute' else existing_hashes.get(file_location)
                if hash_sum:
                    hash_reused += 1
                else:
                    hash_sum = calculate_hashsum(file_path=file_path, transient=True)
                    hash_computed += 1
                if prev_mode == 'skip':
                    prev_version = None
                elif prev_mode == 'recompute':
                    prev_version = get_prev_by_name(
                        file_name=name,
                        object_spec=object_spec,
                        portal=portal_norm
                    )
                    prev_queried += 1
                else:
                    prev_version = existing_prev.get(file_location)
                    if prev_version:
                        prev_reused += 1
                    elif prev_mode == 'auto':
                        prev_version = get_prev_by_name(
                            file_name=name,
                            object_spec=object_spec,
                            portal=portal_norm
                        )
                        prev_queried += 1
                meta = extract_cities_upload_meta(
                    file_path=file_path,
                    file_name=name,
                    description=description,
                    for_station=for_station
                )
                if not prev_version:
                    prev_empty += 1
                rows.append([
                    name,
                    file_location,
                    hash_sum,
                    spec_label,
                    object_spec,
                    prev_version,
                    meta['Level'],
                    meta['title'],
                    meta['created'],
                    meta['startCov'],
                    meta['stopCov'],
                    meta['resolution'],
                    meta['spatialCoverage'],
                    meta['coverageURI'],
                    meta['samplingHeight'],
                    meta['forStation'],
                    meta['creator'],
                    meta['creatorURI'],
                    meta['contributors'],
                    meta['contributorURI'],
                    meta['hostOrganisation'],
                    meta['hostOrganizationURI'],
                    meta['keywords'],
                    meta['licenseName'],
                    meta['licenseUrl'],
                    meta['abstract/description '],
                    meta['comment'],
                    meta['submitterID'],
                    meta['landingPageURI'],
                    meta['doiURI'],
                    meta['variablesToIngest'],
                    meta['documentation'],
                    meta['documentationURI'],
                ])
        except Exception:
            spinner.fail('Metadata extraction failed')
            raise
        spinner.succeed('Metadata extraction completed')

        if hash_computed > 0:
            typer.echo()
        typer.echo('Writing workbook...')
        wb = Workbook()
        ws_envri = wb.active
        ws_envri.title = 'envri_info'
        ws_envri.append(['portal'])
        ws_envri.append([portal_input])
        ws_meta = wb.create_sheet('upload_meta')
        ws_meta.append([
            'fileName',
            'fileLocation',
            'hashSum',
            'type',
            'objectSpecification',
            'isNextVersionOf',
            'Level',
            'title',
            'created',
            'startCov',
            'stopCov',
            'resolution',
            'spatialCoverage',
            'coverageURI',
            'samplingHeight',
            'forStation',
            'creator',
            'creatorURI',
            'contributors',
            'contributorURI',
            'hostOrganisation',
            'hostOrganizationURI',
            'keywords',
            'licenseName',
            'licenseUrl',
            'abstract/description ',
            'comment',
            'submitterID',
            'landingPageURI',
            'doiURI',
            'variablesToIngest',
            'documentation',
            'documentationURI',
        ])
        for row in rows:
            ws_meta.append(row)
        wb.save(output)
        typer.echo(f'Created {output} with {len(files)} file names')
        typer.echo(
            f'Summary: hash reused={hash_reused}, computed={hash_computed}; '
            f'prev reused={prev_reused}, queried={prev_queried}, empty={prev_empty}'
        )
    except Exception as e:
        run_logger.finish(status='error', error=str(e))
        raise
    run_logger.finish(status='ok')


def make_spatial_box(lat_min: float, lat_max: float, lon_min: float, lon_max: float) -> dict[str, Any]:
    return {
        '_type': 'LatLonBox',
        'geo': {
            'coordinates': [[
                [lon_min, lat_min],
                [lon_min, lat_max],
                [lon_max, lat_max],
                [lon_max, lat_min],
                [lon_min, lat_min]
            ]],
            'type': 'Polygon'
        },
        'max': {'lat': lat_max, 'lon': lon_max},
        'min': {'lat': lat_min, 'lon': lon_min}
    }


def build_static_cities_meta(description: str, for_station: str) -> dict[str, str]:
    return {
        'Level': '3',
        'resolution': 'half-hourly',
        'spatialCoverage': '',
        'samplingHeight': '',
        'forStation': for_station,
        'creator': 'Betty Molinier',
        'creatorURI': 'https://citymeta.icos-cp.eu/resources/people/Betty_Molinier',
        'contributors': 'Betty Molinier, Natascha Kljun',
        'contributorURI': json.dumps([
            'https://citymeta.icos-cp.eu/resources/people/Betty_Molinier',
            'https://citymeta.icos-cp.eu/resources/people/Natascha_Kljun'
        ]),
        'hostOrganisation': ORG_DISPLAY_NAMES[ORG_LU_CITIES],
        'hostOrganizationURI': ORG_LU_CITIES,
        'keywords': json.dumps([
            'Flux footprints',
            'atmospheric modelling',
            'urban flux',
            'ICOS Cities'
        ]),
        'licenseName': 'ICOS CCBY4 Data Licence',
        'licenseUrl': 'http://meta.icos-cp.eu/ontologies/cpmeta/icosLicence',
        'abstract/description ': description,
        'comment': (
            'In this version, the axis definition follows the netCDF C API requirements. '
            'Please note that the netCDF C API (this file) interprets data as row major, '
            'and consequently MATLAB users must transpose it.'
        ),
        'submitterID': 'CP',
        'landingPageURI': '',
        'doiURI': '',
        'variablesToIngest': '',
        'documentation': '',
        'documentationURI': 'https://citymeta.icos-cp.eu/objects/ylBBo5HL8RztF6kb0bHclRQd',
    }


def to_landing_page_url(data_url: str) -> str:
    parsed = urlparse(data_url)
    host = parsed.netloc
    if 'data' in host:
        host = host.replace('data', 'meta', 1)
    return urlunparse(parsed._replace(netloc=host))


def extract_cities_upload_meta(
        file_path: Path,
        file_name: str,
        description: str,
        for_station: str
) -> dict[str, str]:
    meta = {
        'title': '',
        'created': '',
        'startCov': '',
        'stopCov': '',
        'coverageURI': '',
    }
    meta.update(build_static_cities_meta(description=description, for_station=for_station))
    ds = None
    try:
        ds = xarray.open_dataset(file_path)
        date = dt.strptime(
            re.split(r'[_\\.]', file_name)[2],
            '%y%m%d'
        ).strftime('%Y-%m-%d')
        meta['title'] = ds.Title.replace('Daily', 'Diurnal') + f' - {date}'
        d_date = ds.Date_Created if 'Date_Created' in ds.attrs else ds.Creation_Date
        meta['created'] = dt.strptime(d_date, '%d-%b-%Y').strftime('%Y-%m-%dT11:00:00Z')
        start_raw = '20' + str(ds.timestep[0].item())
        stop_raw = '20' + str(ds.timestep[-1].item())
        meta['startCov'] = dt.strptime(start_raw, '%Y%m%d%H%M').strftime('%Y-%m-%dT%H:%M:%SZ')
        meta['stopCov'] = dt.strptime(stop_raw, '%Y%m%d%H%M').strftime('%Y-%m-%dT%H:%M:%SZ')
        tm_proj = pyproj.CRS(ds.crs_projection4)
        wgs84_proj = pyproj.CRS('EPSG:4326')
        transformer = pyproj.Transformer.from_crs(
            tm_proj,
            wgs84_proj,
            always_xy=True
        )
        lon, lat = transformer.transform(
            np.array(ds.x, dtype=np.float32),
            np.array(ds.y, dtype=np.float32)
        )
        meta['coverageURI'] = json.dumps(
            make_spatial_box(
                float(lat.min()),
                float(lat.max()),
                float(lon.min()),
                float(lon.max())
            )
        )
    except Exception as e:
        typer.echo(f'Warning: could not extract cities metadata for {file_name}: {e}')
    finally:
        if ds is not None:
            ds.close()
    return meta


def _nan_to_none(obj):
    """Recursively replace float NaN values with None so the payload is JSON-serializable."""
    if isinstance(obj, dict):
        return {key: _nan_to_none(value) for key, value in obj.items()}
    if isinstance(obj, list):
        return [_nan_to_none(item) for item in obj]
    if isinstance(obj, float) and np.isnan(obj):
        return None
    return obj


def make_json(meta: Series):
    description = (
        meta['abstract/description ']
        if 'abstract/description ' in meta
        else meta.get('abstract/description', '')
    )
    spatial_raw = meta.get('coverageURI')
    spatial = None
    if not pd.isna(spatial_raw):
        if isinstance(spatial_raw, str):
            stripped = spatial_raw.strip()
            if stripped.startswith(('http://', 'https://')):
                spatial = stripped
            else:
                spatial = json.loads(stripped)
        else:
            spatial = spatial_raw
    documentation = None
    documentation_uri = meta.get('documentationURI')
    if not pd.isna(documentation_uri):
        doc_raw = str(documentation_uri).strip()
        if doc_raw:
            documentation = doc_raw.rstrip('/').split('/')[-1]
    hash_sum = (
        None
        if pd.isna(meta.get('hashSum'))
        else str(meta.get('hashSum')).strip()
    )
    if not hash_sum:
        data_path = Path(meta['fileLocation']) / meta['fileName']
        if data_path.exists():
            hash_sum = calculate_hashsum(file_path=data_path)
        else:
            typer.echo(f'Hash skipped (data file not found): {data_path}')
    json_meta = dict({
        'fileName': meta['fileName'],
        'hashSum': hash_sum,
        'isNextVersionOf':
            None if pd.isna(meta['isNextVersionOf']) else meta['isNextVersionOf'],
        'preExistingDoi': None if pd.isna(meta['doiURI']) else meta['doiURI'],
        'objectSpecification': meta['objectSpecification'],
        'references': {
            'keywords': json.loads(meta['keywords']),
            'licence': meta['licenseUrl'],
            'autodeprecateSameFilenameObjects': False,
            'duplicateFilenameAllowed': True,
        },
        'specificInfo': {
            'title': meta['title'],
            'description': description,
            'spatial': spatial,
            'temporal': {
                'interval': {
                    'start': meta['startCov'],
                    'stop': meta['stopCov'],
                },
                'resolution': None if pd.isna(meta.get('resolution')) else meta.get('resolution'),
            },
            'forStation': (
                None
                if pd.isna(meta.get('forStation')) or not str(meta.get('forStation')).strip()
                else meta.get('forStation')
            ),
            'production': {
                'creator': meta['creatorURI'],
                'contributors': json.loads(meta['contributorURI']),
                'hostOrganization': meta['hostOrganizationURI'],
                    'comment':
                        None if pd.isna(meta['comment']) else meta['comment'],
                    'sources': [],
                    'documentation': documentation,
                'creationDate': meta['created'],
            },
            'variables': (
                None
                if pd.isna(meta.get('variablesToIngest'))
                or not meta.get('variablesToIngest')
                else json.loads(meta['variablesToIngest'])
            )
        },
        'submitterId': meta['submitterID'],
    })
    return _nan_to_none(json_meta)


def upload_meta(meta_json: dict[str, Any], envri_conf: EnvriConfig) -> tuple[str, str]:
    """Upload metadata package to specified portal."""
    typer.echo(f'Uploading metadata for: {meta_json["fileName"]}', nl=False)
    resp = requests.post(
        url=envri_conf.meta_url,
        json=meta_json,
        cookies=get_cookie_jar()
    )
    response_url = resp.text.strip()
    landing_url = to_landing_page_url(response_url)
    if resp.status_code == 200:
        typer.echo(
            f' -> {landing_url} ({resp.status_code}) OK'
        )
    else:
        typer.echo(f' ({resp.status_code}) FAILED')
        typer.echo(resp.text)
        raise typer.Exit(code=1)
    return response_url, landing_url


def upload_data(file_path: str | Path, data_url: str) -> None:
    """Upload data file to specified portal."""
    file_path = Path(file_path)
    typer.echo(f'Uploading data for: {file_path}', nl=False)
    resp = requests.put(
        url=data_url,
        data=open(file=file_path, mode='rb'),
        cookies=get_cookie_jar(),
        headers={'Content-Type': 'application/octet-stream'},
    )
    if resp.status_code == 200:
        typer.echo(f' -> {resp.text} ({resp.status_code}) OK')
    else:
        typer.echo(f' ({resp.status_code}) FAILED')
        typer.echo(resp.text)
        raise typer.Exit(code=1)
    return


if __name__ == "__main__":
    app()
